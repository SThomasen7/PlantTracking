#!/home/stasen/Desktop/plantas/env/bin/python
from ultralytics import YOLO
import cv2
from optparse import OptionParser
import os
import pathlib
from openpyxl import Workbook, load_workbook

parser = OptionParser()

parser.add_option(
    "-m", "--model",
    dest="model",
    help="Path to YOLO model weights",
    metavar="MODEL"
)

parser.add_option(
    "-v", "--video",
    dest="video",
    help="Path to input video",
    metavar="FILE"
)

parser.add_option(
    "-p", "--multiprocess",
    dest="multiprocess",
    help="Path to directory",
    metavar="DIR"
)

parser.add_option(
    "-w", "--excel-workbook",
    dest="workbook",
    help="Excel spreadsheet to write to"
)

parser.add_option(
    "-o", "--output-dir",
    dest="output",
    help="Excel spreadsheet to write to"
)

(options, args) = parser.parse_args()

## excel workbook
output_excel = options.workbook
if output_excel is None:
    output_excel = "counts.xlsx"

output_dir = options.output
if output_dir is None:
    output_dir = os.getcwd()
print(output_dir)

def track_video(model, video):
    cap = cv2.VideoCapture(video)

    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps = cap.get(cv2.CAP_PROP_FPS)

    os.makedirs(output_dir, exist_ok=True)

    base_file = os.path.basename(video)
    out_path = os.path.join(output_dir, f"{base_file}_tracked.mp4")
    print(out_path)
    out = cv2.VideoWriter(
        out_path,
        cv2.VideoWriter_fourcc(*"mp4v"),
        fps,
        (w, h)
    )

    results = model.track(
        source=video,
        conf=0.4,
        iou=0.5,
        tracker="bytetrack.yaml",
        stream=True
    )

    ## Iterate over the frames and get the count of unique ids
    seen_ids = dict()
    for r in results:
        frame = r.plot()

        if r.boxes.id is not None:
            ids = r.boxes.id.cpu().numpy()
            for i in ids:
                if int(i) not in seen_ids.keys():
                    seen_ids[int(i)] = 0
                seen_ids[int(i)] += 1

        total = 0
        for key, val in seen_ids.items():
            if val > 5:
                total += 1

        cv2.putText(
            frame,
            f"Conteo: {total}",
            (40, 60),
            cv2.FONT_HERSHEY_SIMPLEX,
            1.0,
            (100, 255, 100),
            2
        )

        out.write(frame)

    # write the output
    out_path = os.path.join(output_dir, output_excel)
    if os.path.exists(out_path):
        wb = load_workbook(out_path)
        ws = wb.active
    else:
        # if we are making a new workbook add table headers.
        wb = Workbook()
        ws = wb.active
        ws.append(["Video", "Conteo"])

    ws.append([video, total])
    wb.save(out_path)

    out.release()


if __name__ == "__main__":

    if options.model is None:
        model_path = "runs/detect/train2/weights/best.pt"
    else:
        model_path = options.model

    if options.multiprocess is not None:

        files = os.listdir(options.multiprocess)
        video_path = pathlib.Path(options.multiprocess)

        for video in video_path.iterdir():
            print(f"Detecting video: {video}")
            print(f"Model: {model_path}")
            
            model = YOLO(model_path)
            track_video(model, str(video))

    else:
        if options.video is None:
            video_path = "/home/stasen/Desktop/plantas/videos/linea_841.mp4"
        else:
            video_path = options.video

        print(f"Detecting video: {video_path}")
        print(f"Model: {model_path}")
        
        model = YOLO(model_path)
        track_video(model, video_path)

